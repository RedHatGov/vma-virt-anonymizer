#!/usr/bin/env python3
"""
RVTools Anonymizer - Python Validation and Testing Script

This script provides:
1. Validation of anonymized RVTools exports
2. A Python-based anonymization alternative (for testing)
3. Analysis of sensitive data in RVTools exports

Usage:
    python validate_anonymization.py --analyze <rvtools_file.xlsx>
    python validate_anonymization.py --anonymize <rvtools_file.xlsx>
    python validate_anonymization.py --validate <original.xlsx> <anonymized.xlsx>
"""

import argparse
import re
import sys
from pathlib import Path
from collections import defaultdict
from typing import Dict, Set, Optional, Tuple

try:
    import openpyxl
    from openpyxl import Workbook
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# Columns that contain sensitive data, mapped to their data type
SENSITIVE_COLUMNS = {
    'VM': 'vm',
    'Name': 'name',  # Can be datastore, cluster, etc. depending on sheet
    'Host': 'host',
    'Cluster': 'cluster',
    'Datacenter': 'datacenter',
    'DNS Name': 'dns',
    'Domain': 'domain',
    'DNS Search Order': 'domain',
    'DNS Servers': 'ip',
    'Folder': 'folder',
    'vApp': 'folder',
    'Resource pool': 'folder',
    'Path': 'path',
    'Network': 'network',
    'Portgroup': 'network',
    'VI SDK Server': 'ip',
    'Primary IP Address': 'ip',
    'Annotation': 'annotation',
    'Notes': 'annotation',
}


class RVToolsAnonymizer:
    """Anonymizes RVTools export files while maintaining consistency."""
    
    def __init__(self):
        self.mappings: Dict[str, Dict[str, str]] = defaultdict(dict)
        self.counters: Dict[str, int] = defaultdict(int)
        self.prefixes = {
            'vm': 'VM-',
            'host': 'HOST-',
            'cluster': 'CLUSTER-',
            'datacenter': 'DC-',
            'datastore': 'DS-',
            'network': 'NET-',
            'folder': 'FOLDER-',
            'domain': 'domain',
            'ip': '10.0.',
            'name': 'NAME-',
            'annotation': '',
        }
    
    def get_anonymized_value(self, original: str, data_type: str) -> str:
        """Get or create an anonymized value for the given original."""
        if not original or str(original).strip() == '':
            return ''
        
        original = str(original)
        
        # Check if already mapped
        if original in self.mappings[data_type]:
            return self.mappings[data_type][original]
        
        # Generate new anonymized value
        self.counters[data_type] += 1
        counter = self.counters[data_type]
        prefix = self.prefixes.get(data_type, 'ANON-')
        
        if data_type == 'ip':
            new_value = f"10.0.{counter // 256}.{counter % 256}"
        elif data_type == 'domain':
            new_value = f"domain{counter}.local"
        elif data_type == 'annotation':
            new_value = ''  # Clear annotations
        else:
            new_value = f"{prefix}{counter:04d}"
        
        self.mappings[data_type][original] = new_value
        return new_value
    
    def anonymize_dns_name(self, dns_name: str) -> str:
        """Anonymize a DNS name, handling both hostname and domain parts."""
        if not dns_name or '.' not in str(dns_name):
            return self.get_anonymized_value(dns_name, 'vm')
        
        parts = str(dns_name).split('.', 1)
        hostname = parts[0]
        domain = parts[1] if len(parts) > 1 else ''
        
        anon_host = self.get_anonymized_value(hostname, 'vm')
        anon_domain = self.get_anonymized_value(domain, 'domain') if domain else ''
        
        return f"{anon_host}.{anon_domain}" if anon_domain else anon_host
    
    def anonymize_path(self, path: str) -> str:
        """Anonymize datastore paths like [DATASTORE] VM/VM.vmx
        Also anonymizes VM folder names and filenames in the path."""
        if not path:
            return ''
        
        path = str(path)
        
        # Handle [DATASTORE] VM_FOLDER/VM_FILE.ext paths
        match = re.match(r'\[([^\]]+)\]\s*(.*)', path)
        if match:
            ds_name = match.group(1)
            rest_of_path = match.group(2).strip()
            anon_ds = self.get_anonymized_value(ds_name, 'datastore')
            
            # Check if there's a folder/file structure
            if '/' in rest_of_path:
                parts = rest_of_path.split('/', 1)
                vm_folder = parts[0]
                filename = parts[1] if len(parts) > 1 else ''
                
                # Anonymize the VM folder name
                anon_vm_folder = self.get_anonymized_value(vm_folder, 'vm')
                
                # Replace original VM name in filename with anonymized version
                if filename and vm_folder:
                    anon_filename = filename.replace(vm_folder, anon_vm_folder)
                else:
                    anon_filename = filename
                
                return f"[{anon_ds}] {anon_vm_folder}/{anon_filename}"
            else:
                # Just datastore and possibly a filename
                return f"[{anon_ds}] {rest_of_path}"
        
        return self.get_anonymized_value(path, 'folder')
    
    def process_cell(self, value, column_name: str, sheet_name: str) -> str:
        """Process a cell value based on its column type."""
        if value is None or str(value).strip() == '':
            return value
        
        str_value = str(value)
        data_type = SENSITIVE_COLUMNS.get(column_name)
        
        if not data_type:
            return value
        
        # Special handling based on column type
        if column_name == 'DNS Name':
            return self.anonymize_dns_name(str_value)
        elif column_name == 'Host' and '.' in str_value:
            return self.anonymize_dns_name(str_value)
        elif column_name == 'Path':
            return self.anonymize_path(str_value)
        elif column_name == 'Name':
            # "Name" varies by sheet
            if sheet_name in ['vDatastore']:
                return self.get_anonymized_value(str_value, 'datastore')
            elif sheet_name in ['vCluster']:
                return self.get_anonymized_value(str_value, 'cluster')
            else:
                return self.get_anonymized_value(str_value, 'name')
        else:
            return self.get_anonymized_value(str_value, data_type)
    
    def anonymize_workbook(self, input_path: str, output_path: Optional[str] = None) -> str:
        """Anonymize an entire RVTools workbook."""
        input_path = Path(input_path)
        if output_path is None:
            output_path = input_path.parent / f"{input_path.stem}_anonymized{input_path.suffix}"
        else:
            output_path = Path(output_path)
        
        print(f"Loading workbook: {input_path}")
        wb = openpyxl.load_workbook(input_path)
        
        total_sheets = len(wb.sheetnames)
        for idx, sheet_name in enumerate(wb.sheetnames, 1):
            print(f"Processing sheet {idx}/{total_sheets}: {sheet_name}")
            ws = wb[sheet_name]
            self._anonymize_sheet(ws, sheet_name)
        
        print(f"Saving anonymized workbook: {output_path}")
        wb.save(output_path)
        
        return str(output_path)
    
    def _anonymize_sheet(self, ws, sheet_name: str):
        """Anonymize a single worksheet."""
        # Get headers from first row
        headers = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                headers[col] = str(header)
        
        # Process each data row
        for row in range(2, ws.max_row + 1):
            for col, header in headers.items():
                if header in SENSITIVE_COLUMNS:
                    cell = ws.cell(row=row, column=col)
                    if cell.value is not None:
                        cell.value = self.process_cell(cell.value, header, sheet_name)
    
    def export_mappings(self, output_path: str):
        """Export all mappings to an Excel file."""
        wb = Workbook()
        
        # Remove default sheet
        default_sheet = wb.active
        
        for data_type, mapping in self.mappings.items():
            if not mapping:
                continue
            
            ws = wb.create_sheet(title=f"{data_type}_mappings"[:31])
            ws.cell(row=1, column=1, value="Original Value")
            ws.cell(row=1, column=2, value="Anonymized Value")
            
            for row, (original, anonymized) in enumerate(mapping.items(), start=2):
                ws.cell(row=row, column=1, value=original)
                ws.cell(row=row, column=2, value=anonymized)
        
        # Remove default sheet if we created others
        if len(wb.sheetnames) > 1:
            wb.remove(default_sheet)
        
        wb.save(output_path)
        print(f"Mappings exported to: {output_path}")


def analyze_rvtools_file(file_path: str):
    """Analyze an RVTools file for sensitive data."""
    print(f"\n{'='*60}")
    print(f"Analyzing: {file_path}")
    print('='*60)
    
    wb = openpyxl.load_workbook(file_path, read_only=True)
    
    sensitive_data: Dict[str, Set[str]] = defaultdict(set)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Get headers
        headers = {}
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            for col, header in enumerate(row, 1):
                if header and str(header) in SENSITIVE_COLUMNS:
                    headers[col-1] = str(header)
            break
        
        if not headers:
            continue
        
        # Sample values from first 100 rows
        row_count = 0
        for row in ws.iter_rows(min_row=2, max_row=101, values_only=True):
            row_count += 1
            for col, header in headers.items():
                if col < len(row) and row[col]:
                    sensitive_data[header].add(str(row[col])[:100])
    
    print(f"\nSheets analyzed: {len(wb.sheetnames)}")
    print(f"\nSensitive columns found:")
    
    for col_name, values in sorted(sensitive_data.items()):
        print(f"\n  {col_name} ({len(values)} unique values sampled):")
        for val in list(values)[:5]:
            print(f"    - {val}")
        if len(values) > 5:
            print(f"    ... and {len(values) - 5} more")


def validate_anonymization(original_path: str, anonymized_path: str):
    """Validate that anonymization was performed correctly."""
    print(f"\n{'='*60}")
    print("Validating Anonymization")
    print('='*60)
    
    orig_wb = openpyxl.load_workbook(original_path, read_only=True)
    anon_wb = openpyxl.load_workbook(anonymized_path, read_only=True)
    
    issues = []
    stats = {
        'sheets_checked': 0,
        'cells_checked': 0,
        'cells_anonymized': 0,
        'original_values_leaked': 0,
    }
    
    # Collect all original sensitive values
    original_values: Set[str] = set()
    
    for sheet_name in orig_wb.sheetnames:
        ws = orig_wb[sheet_name]
        headers = {}
        
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            for col, header in enumerate(row):
                if header and str(header) in SENSITIVE_COLUMNS:
                    headers[col] = str(header)
            break
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            for col in headers:
                if col < len(row) and row[col]:
                    original_values.add(str(row[col]).lower())
    
    print(f"Collected {len(original_values)} unique sensitive values from original")
    
    # Check anonymized file for leaked values
    for sheet_name in anon_wb.sheetnames:
        stats['sheets_checked'] += 1
        ws = anon_wb[sheet_name]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            for cell in row:
                if cell:
                    stats['cells_checked'] += 1
                    cell_str = str(cell).lower()
                    
                    # Check if any original value appears in this cell
                    for orig_val in original_values:
                        if len(orig_val) > 3 and orig_val in cell_str:
                            stats['original_values_leaked'] += 1
                            if len(issues) < 20:
                                issues.append(f"Potential leak in {sheet_name}: '{cell}' contains '{orig_val}'")
    
    print(f"\nValidation Results:")
    print(f"  Sheets checked: {stats['sheets_checked']}")
    print(f"  Cells checked: {stats['cells_checked']}")
    print(f"  Potential leaks found: {stats['original_values_leaked']}")
    
    if issues:
        print(f"\nIssues found:")
        for issue in issues:
            print(f"  - {issue}")
    else:
        print(f"\n✓ No obvious data leaks detected!")


def main():
    parser = argparse.ArgumentParser(
        description='RVTools Anonymizer - Validation and Testing Tool',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  Analyze a file for sensitive data:
    python validate_anonymization.py --analyze data/rvtools_export.xlsx

  Anonymize a file:
    python validate_anonymization.py --anonymize data/rvtools_export.xlsx

  Validate anonymization:
    python validate_anonymization.py --validate original.xlsx anonymized.xlsx
        """
    )
    
    parser.add_argument('--analyze', metavar='FILE',
                        help='Analyze an RVTools file for sensitive data')
    parser.add_argument('--anonymize', metavar='FILE',
                        help='Anonymize an RVTools file (creates _anonymized copy)')
    parser.add_argument('--validate', nargs=2, metavar=('ORIGINAL', 'ANONYMIZED'),
                        help='Validate that anonymization was performed correctly')
    parser.add_argument('--output', '-o', metavar='FILE',
                        help='Output path for anonymized file')
    parser.add_argument('--export-mappings', action='store_true',
                        help='Export value mappings to a separate file')
    
    args = parser.parse_args()
    
    if args.analyze:
        analyze_rvtools_file(args.analyze)
    
    elif args.anonymize:
        anonymizer = RVToolsAnonymizer()
        output_path = anonymizer.anonymize_workbook(args.anonymize, args.output)
        print(f"\n✓ Anonymization complete: {output_path}")
        
        if args.export_mappings:
            mapping_path = Path(output_path).parent / "anonymization_mappings.xlsx"
            anonymizer.export_mappings(str(mapping_path))
    
    elif args.validate:
        validate_anonymization(args.validate[0], args.validate[1])
    
    else:
        parser.print_help()


if __name__ == '__main__':
    main()
